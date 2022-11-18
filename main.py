from openpyxl import load_workbook

workbook = load_workbook(filename="prc.xlsm")
print(f"Worksheet names: {workbook.sheetnames}")
sheet = workbook.active
print(sheet)
print(f"The title of the Worksheet is: {sheet.title}")